#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Autor:Luisa Fernanda Buriticá Ruíz
# Fecha:22 junio de 2023
# Conctacto:fernanda.buritica@udea.edu.co
# Titulo: Funciones de lectura de archivos AMSC
###############################################################################
# 1. LIBRERIAS
# 1.1 Librerías importadas
import csv                       # Funcionalidades para trabajar con archivos CSV.
import glob                      # Identificar archivos que coincidan con un patrón específico.
import matplotlib.pyplot as plt  # Generación de gráficos
import numpy             as np   # Herramienta matemática.
import pandas            as pd   # Procesar datos.
import os                        # Leer archivos dentro de una carpeta.
import re                        # Trabajar con expresiones regulares
import time

# 1.2 Modulos importados
from dateutil.parser import ParserError
from datetime import datetime
from openpyxl import load_workbook #Cargar y manipular archivos de hojas de cálculo en formato XLSX (Excel)
from tqdm import tqdm       # Tiempo de ejecución
from unidecode import unidecode  # Manejar y normalizar cadenas de texto que contienen caracteres

###############################################################################
# 2. FUNCIONES
def graficoST(df,estacion):
    dicc_variables={
        "Bar":["Presión","Hpa"],
        "Out_Temp":["Temperatura","°C"],
        "Wind_Speed":["Velocidad del viento","m/seg"],
        "Wind_Dir":["Dirección del viento","%"],
        "Out_Hum":["Humedad Exterior","%"],
        "Rain_Rate":["Precipitación","cm/hr"],
        }  
    
    for variable, info in dicc_variables.items():
        nombre_variable = variable
        descripcion = info[0]
        unidad = info[1]

        fig = plt.figure(figsize=(10,5))
        ax = fig.add_subplot(111)
        plt.title("Serie de tiempo: {} ".format(descripcion))
        plt.plot(df["fecha"],df[variable],label=estacion)
        plt.ylabel("{} [{}]".format(descripcion,unidad),fontsize=12)
        plt.xlabel("Tiempo (año-mes)",fontsize=12)
        plt.grid()
        plt.minorticks_on()

            #####################################################
def resumen_estacion(Estacion,path_in3,columnas,path_out1):
    archivos=[]
    C1 = os.listdir(path_in3)
    print("PASO 1")
    for i in tqdm(range(len(C1))):
        print("PASO 2 -",len(C1))
        estacion=unidecode(C1[i].replace("DatosMeteo","").lower().strip())
        #Lectura de las carpetas por estación
        if estacion==Estacion:# Filtro para seleccionar la estación
            print("#######################################\n")
            print(estacion)
            C2=glob.glob(path_in3+C1[i]+"*/A20*")
            #Se recorren las carpetas
            v=pd.DataFrame() #Almacenar todos los csv
            for ii in range(len(C2)):
                C3=os.listdir(C2[ii])
                for iii in range(len(C3)):
                    try:
                        m=C3[iii].split(".")
                        if "swp" in m: continue
                        datos=pd.read_csv(C2[ii]+"/"+C3[iii],usecols=columnas)
                    
                    except pd.errors.EmptyDataError as e:
                        # Guardar información sobre el error
                        print(f"Ocurrió un error: {e} - Sin Columnas")
                        #print(f"Información del error: {error_info}")
                        archivos.append(C3[iii])
                    
                    except pd.errors.ParserError as e:
                        # Guardar información sobre el error
                        print(f"Ocurrió un error: {e} - Errore de Parseo")
                        #print(f"Información del error: {error_info}")
                        archivos.append(C3[iii])
                        with open(C3[iii], newline='') as csvfile:
                            df = csv.reader(csvfile)
                            next(df) # Omitir la primera fila si es un encabezado
                            datos=[]
                            #print(1)
                            for iv, fila in enumerate(df, start=2):
                                m=f'{fila}'.replace("'","").replace("[","").replace("]","").split(",")
                                if len(m)==42:
                                    datos.append(m)
                                if len(m)!=42:
                                    print("En la fila ",iv," existe un error de parseo")
                                    print("")
                            datos=pd.DataFrame(datos)
                            print(datos)
                    
                    
                    v=pd.concat([v,datos],axis=0)
                    v=v.reset_index(drop=True)
            #Verificación de nombre de columnas
            m=v.columns
            print(2*"#","\n",m)
            v.columns=["fecha","Bar","Out_Temp","Wind_Speed","Wind_Dir","Out_Hum","Rain_Rate","UV","Sol_Rad"]
            m=v.columns
            print(2*"#",m,"\n")
            #Fecha
            try:
                v['fecha'] = pd.to_datetime(v['fecha'])
                print("\n fecha inicial: ", v.fecha.min(),
                      "\n fecha final: ", v.fecha.max())
            except ParserError :
                start_time = time.time()
                v = v.dropna(subset=['fecha'])
                print("Comienza proceso para procesar fecha, puede tardar un tiempo")
                def convertir_fecha(texto):
                    try:
                        fecha_convertida = pd.to_datetime(texto, format='%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        fecha_convertida = pd.to_datetime(texto, format='%Y:%m:%d-%H:%M:%S', errors='coerce')
                    return fecha_convertida
                v['fecha'] = v['fecha'].apply(convertir_fecha)
                v['fecha'] = v['fecha'] 
                end_time = time.time()
                execution_time = end_time - start_time
                print("\n fecha inicial: ", v.fecha.min(),
                      "\n fecha final: ", v.fecha.max(),execution_time)
            except ValueError as e :  
                print(e)
                position_start = str(e).find("position") + len("position") + 1
                position_end = str(e).find(".", position_start)
                position = int(str(e)[position_start:position_end])
                print("La posición del error es:", position)
                
                v=v.drop([int(position)],axis=0)
                v=v.reset_index(drop=True)
                v['fecha'] = pd.to_datetime(v['fecha']) 
                            
            #Variables
            v["Rain_Rate"]=v['Rain_Rate']*0.2/60. #in units of cm/hr
            if Estacion in ["caucasia","santafe"]:
                v['Bar'] = (v['Bar']/1000.*(3386.389/100.0))
            else:
                v['Bar'] = (v['Bar']/1000.*(3386.389/100.0))
            v['Out_Temp'] = ( v['Out_Temp']/10. - 32.) * (5.0/9.0)
            v.loc[(v['Out_Temp'] > 50) | (v['Out_Temp'] < -15)] = np.nan
            #Procesamiento final
            v.sort_values("fecha")
            v = v.dropna(subset=['fecha'])
            diferencia = v['fecha'].diff()
            frecuencia_comun = diferencia.mode()[0]
            v = v.resample(frecuencia_comun, on='fecha').mean()
            v=v.reset_index()
            v.to_csv(path_out1+estacion+".csv")
            #Visualizacion
            #graficoST(v,estacion)
            print("\n #######################################")
            #####################################################
def resumen_estacion_andes(Estacion,path_in3,columnas_xlsx,path_out1):
    v=pd.DataFrame()
    C4 = glob.glob(path_in3+"/*DatosMe*")
    for i in range(len(C4)):
        print(40*"#")
        estacion=unidecode(C4[i].split("/")[7].replace("DatosMeteo","").lower().strip())
        
        #Lectura de las carpetas por estación
        if estacion==Estacion:# Filtro para seleccionar la estación
            print("\n",estacion,"\n")
            C5=glob.glob(C4[i]+"/*xlsx")
            
            print(20*"#","\n","# ARHIVOS XLSX# \n",20*"#")
            for ii in tqdm(range(len(C5))):
                wb = load_workbook(filename=C5[ii])
                nombres_hojas = wb.sheetnames
                for iii in tqdm(nombres_hojas):
                    # Obtener la hoja actual por su nombre
                    hoja = wb[iii]
                    df = pd.DataFrame(hoja.values)
                    df=df.drop([0,1,2,3,4],axis=0).reset_index(drop=True)
                    c=df.columns
                    if len(c)==21:
                        df.columns=columnas_xlsx
                    if len(c)==20:
                        df.columns=columnas_xlsx[:20]
                        df["Solar_Energy"]=np.nan
                    for index, row in df.iterrows():
                        try: 
                            datetime.strptime(row["Date"], "%d/%m/%y")
                        except ValueError:
                            df=df.drop([index],axis=0)
                        except TypeError as error:
                            print("Ocurrió un error de tipo:", error)
                    
                    v=pd.concat([v,df],axis=0)
            C6=glob.glob(C4[i]+"/*csv")
            print(20*"#","\n","# ARHIVOS CSV# \n",20*"#")
            for iv in tqdm(range(len(C6))):
                df=pd.read_csv(C6[iv])
                df.columns=columnas_xlsx
                for index, row in tqdm(df.iterrows()):
                    try: 
                        datetime.strptime(row["Date"], "%d/%m/%y")
                    except ValueError:
                        df=df.drop([index],axis=0)
                v=pd.concat([v,df],axis=0)
        #######################################
            v=v.reset_index(drop=True)
            v["fecha"]=np.nan
            count=0
            for index, row in tqdm(v.iterrows()):
                tipo_dato = type(row["Date"])
                if tipo_dato == datetime:
                    anio=row["Date"].year
                    mes=row["Date"].month
                    dia=row["Date"].day
                    hora=int(row["Time"].split(":")[0])
                    minuto=int(re.sub('[a-zA-Z]', '',row["Time"].split(":")[1]))
                    meridiano= re.sub(r'(\d+|:)', '', row["Time"])
                    if "a" in meridiano:
                        v["fecha"][index]=datetime(anio,mes,dia,hora,minuto)
                    if "p" in meridiano:
                        v["fecha"][index]=datetime(anio,mes,dia,hora,minuto+12)
                if tipo_dato==str:        
                    fecha=row["Date"]+" "+row["Time"]
                    meridiano=re.sub(r"[\d\/:\s]+","",fecha)
                    if  "a" in meridiano: 
                        fecha=re.sub('[a-zA-Z]', '', fecha)
                        fecha=fecha+"am"
                    if  "p" in meridiano:
                        fecha=re.sub('[a-zA-Z]', '', fecha)
                        fecha=fecha+"pm"
                    try:
                        v["fecha"][index]=datetime.strptime(fecha, "%d/%m/%y %I:%M%p")
                    except ValueError:
                        v["fecha"][index]=datetime.strptime(fecha, "%d/%m/%y %I:%M %p")
                        count+=1
            v1 = pd.concat([v.fecha, v["Bar"], v["Out_Temp"], v["Wind_Speed"],
                  v["Wind_Dir"], v["Out_Hum"], v["Rain_Rate"],
                  v["Solar_Energy"], v["Solar_Rad"]], axis=1)
            # Fecha
            try:
                v1['fecha'] = pd.to_datetime(
                    v1['fecha']) 
                print("\n fecha inicial: ", v1.fecha.min(),
                      "\n fecha final: ", v1.fecha.max())
            except ParserError:
                start_time = time.time()
                v1 = v1.dropna(subset=['fecha'])
                print("Comienza proceso para procesar fecha, puede tardar un tiempo")
                
                def convertir_fecha(texto):
                    try:
                        fecha_convertida = pd.to_datetime(
                            texto, format='%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        fecha_convertida = pd.to_datetime(
                            texto, format='%Y:%m:%d-%H:%M:%S', errors='coerce')
                    return fecha_convertida
                v1['fecha'] = v1['fecha'].apply(convertir_fecha)
                v1['fecha'] = v1['fecha'] 
                end_time = time.time()
                execution_time = end_time - start_time
                print("\n fecha inicial: ", v1.fecha.min(),
                      "\n fecha final: ", v1.fecha.max(), execution_time)
            #Variables
            # Correccion de datos
            m=list(v1.columns)
            m.remove("fecha")
            for i in range(len(m)):
                v1[m[i]] = pd.to_numeric(v1[m[i]], errors='coerce')
                v1[m[i]] = v1[m[i]].astype(float)
            
            #Procesamiento final
            v1.sort_values("fecha")
            v1 = v1.dropna(subset=['fecha'])
            diferencia = v1['fecha'].diff()
            frecuencia_comun = diferencia.mode()[0]
            v1 = v1.resample(frecuencia_comun, on='fecha').mean()
            v1=v1.reset_index()
            v1.to_csv(path_out1+estacion+".csv")
            #Visualizacion
            graficoST(v1,estacion)
            print("\n #######################################")
            #####################################################
def resumen_estacion_oriente(Estacion,path_in3,columnas,path_out1):
    v=pd.DataFrame()
    C4 = glob.glob(path_in3+"/*DatosMe*")
    for i in range(len(C4)): 
        estacion=unidecode(C4[i].split("/")[7].replace("DatosMeteo","").lower().strip())
        #print(estacion)
        if estacion==Estacion:
            C8=glob.glob(C4[i]+"*/A20*")
            for ii in tqdm(range(len(C8))):
                C9=glob.glob(C8[ii]+"*/*oldformat*")
                if len(C9)!=0:
                    C9=glob.glob(C8[ii]+"*/*oldformat*")[0]
                    C9_1=glob.glob(C9+"*/*csv*")
                    for iv in range(len(C9_1)):
                        datos=pd.read_csv(C9_1[iv],usecols=columnas,encoding='latin-1')
                        def convertir_fecha(texto):
                            fecha_convertida = datetime.strptime(texto, "%Y:%m:%d-%H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")
                            return fecha_convertida
                        datos['Tiempo Sistema'] = datos['Tiempo Sistema'].apply(convertir_fecha)
                        v=pd.concat([v,datos],axis=0)
                        v=v.reset_index(drop=True)
                        
                C10=glob.glob(C8[ii]+"*/*csv*")
                for iii in range(len(C10)):
                    datos=pd.read_csv(C10[iii],usecols=columnas)
                    v=pd.concat([v,datos],axis=0)
            #Verificación de nombre de columnas
            m=v.columns
            print(2*"#","\n",m)
            v.columns=["fecha","Bar","Out_Temp","Wind_Speed","Wind_Dir","Out_Hum","Rain_Rate","UV","Sol_Rad"]
            m=v.columns
            print(2*"#",m,"\n")
            #Fecha
            try:
                v['fecha'] = pd.to_datetime(v['fecha']) - pd.to_timedelta(5,unit='h')
                print("\n fecha inicial: ", v.fecha.min(),
                      "\n fecha final: ", v.fecha.max())
            except ParserError :
                start_time = time.time()
                v = v.dropna(subset=['fecha'])
                print("Comienza proceso para procesar fecha, puede tardar un tiempo")
                def convertir_fecha(texto):
                    try:
                        fecha_convertida = pd.to_datetime(texto, format='%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        fecha_convertida = pd.to_datetime(texto, format='%Y:%m:%d-%H:%M:%S', errors='coerce')
                    return fecha_convertida
                v['fecha'] = v['fecha'].apply(convertir_fecha)
                v['fecha'] = v['fecha'] - pd.to_timedelta(5, unit='h')
                end_time = time.time()
                execution_time = end_time - start_time
                print("\n fecha inicial: ", v.fecha.min(),
                      "\n fecha final: ", v.fecha.max(),execution_time)
            #Variables
            m=list(v.columns)
            m.remove("fecha")
            for i in range(len(m)):
                v[m[i]] = pd.to_numeric(v[m[i]], errors='coerce')
                v[m[i]] = v[m[i]].astype(float)
            v["Rain_Rate"]=v['Rain_Rate']*0.2/60. #in units of cm/hr
            if Estacion in ["caucasia","santafe"]:
                v['Bar'] = (v['Bar']/1000.*(3386.389/100.0))
            else:
                v['Bar'] = (v['Bar']/1000.*(3386.389/100.0))
            v['Out_Temp'] = ( v['Out_Temp']/10. - 32.) * (5.0/9.0)
            v.loc[(v['Out_Temp'] > 50) | (v['Out_Temp'] < -15)] = np.nan
            #Procesamiento final
            v.sort_values("fecha")
            v = v.dropna(subset=['fecha'])
            diferencia = v['fecha'].diff()
            frecuencia_comun = diferencia.mode()[0]
            v = v.resample(frecuencia_comun, on='fecha').mean()
            v=v.reset_index()
            v.to_csv(path_out1+estacion+".csv")
            #Visualizacion
            graficoST(v,estacion)
            print("\n #######################################")
            #####################################################
def resumen_estacion_pb(Estacion,path_in3,columnas_CSV,path_out1):
    v=pd.DataFrame()
    C4 = glob.glob(path_in3+"/*DatosMe*")

    for i in range(len(C4)): 
        estacion=unidecode(C4[i].split("/")[7].replace("DatosMeteo","").lower().strip())
        #print(estacion)
        if estacion==Estacion:
            C11=glob.glob(C4[i]+"*/*CSV*")
            for ii in range(len(C11)):
                datos=pd.read_csv(C11[ii],usecols=columnas_CSV)
                v=pd.concat([v,datos],axis=0)
            m=v.columns
            print(2*"#","\n",m)
            v.columns=["fecha","Out_Temp","Out_Hum","Bar","Rain_Rate","Wind_Speed","Wind_Dir"]
            m=v.columns
            print(2*"#",m,"\n")
            #Fecha
            try:
                v['fecha'] = pd.to_datetime(v['fecha']) 
                print("\n fecha inicial: ", v.fecha.min(),
                      "\n fecha final: ", v.fecha.max())
            except ParserError :
                start_time = time.time()
                v = v.dropna(subset=['fecha'])
                print("Comienza proceso para procesar fecha, puede tardar un tiempo")
                def convertir_fecha(texto):
                    try:
                        fecha_convertida = pd.to_datetime(texto, format='%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        fecha_convertida = pd.to_datetime(texto, format='%Y:%m:%d-%H:%M:%S', errors='coerce')
                    return fecha_convertida
                v['fecha'] = v['fecha'].apply(convertir_fecha)
                v['fecha'] = v['fecha'] 
                end_time = time.time()
                execution_time = end_time - start_time
                print("\n fecha inicial: ", v.fecha.min(),
                      "\n fecha final: ", v.fecha.max(),execution_time)
            #Variables
            m=list(v.columns)
            m.remove("fecha")
            for i in range(len(m)):
                v[m[i]] = pd.to_numeric(v[m[i]], errors='coerce')
                v[m[i]] = v[m[i]].astype(float)
            
            v["Rain_Rate"]=v['Rain_Rate']*0.2/60. #in units of cm/hr
            if Estacion in ["caucasia","santafe"]:
                v['Bar'] = (v['Bar']/1000.*(3386.389/100.0))
            else:
                v['Bar'] = (v['Bar']/1000.*(3386.389/100.0))
            v['Out_Temp'] = ( v['Out_Temp']/10. - 32.) * (5.0/9.0)
            v.loc[(v['Out_Temp'] > 50) | (v['Out_Temp'] < -15)] = np.nan
            #Procesamiento final
            v.sort_values("fecha")
            v = v.dropna(subset=['fecha'])
            diferencia = v['fecha'].diff()
            frecuencia_comun = diferencia.mode()[0]
            v = v.resample(frecuencia_comun, on='fecha').mean()
            v=v.reset_index()
            v.to_csv(path_out1+estacion+".csv")
            #Visualizacion
            graficoST(v,estacion)
            
            print("\n #######################################")
            #####################################################
def resumen_estacion_sonson(Estacion,path_in3,columnas,path_out1):
    v=pd.DataFrame()
    C4 = glob.glob(path_in3+"/*DatosMe*")
    for i in range(len(C4)): 
        estacion=unidecode(C4[i].split("/")[7].replace("DatosMeteo","").lower().strip())

        if estacion==Estacion:
            C12=glob.glob(C4[i]+"*/A20*")  
            for ii in tqdm(range(len(C12))):
                C13=glob.glob(C12[ii]+"/*.csv*")
                for iii in range(len(C13)):
                    try:
                        datos=pd.read_csv(C13[iii],usecols=columnas)
                        v=pd.concat([v,datos],axis=0)
                        
                    except pd.errors.EmptyDataError:
                        print("EmptyDataError: No columns to parse from file")
                        
                    except ValueError as e:
                        print("#\n",i,ii,iii,"\n#")
                        print(e)
            m=v.columns
            print(2*"#","\n",m)
            v.columns=["fecha","Bar","Out_Temp","Wind_Speed","Wind_Dir","Out_Hum","Rain_Rate","UV","Sol_Rad"]
            m=v.columns
            print(2*"#",m,"\n")

            #Fecha
            try:
                v['fecha'] = pd.to_datetime(v['fecha']) 
                print("\n fecha inicial: ", v.fecha.min(),
                      "\n fecha final: ", v.fecha.max())
            except ParserError :
                start_time = time.time()
                v = v.dropna(subset=['fecha'])
                print("Comienza proceso para procesar fecha, puede tardar un tiempo")
                def convertir_fecha(texto):
                    try:
                        fecha_convertida = pd.to_datetime(texto, format='%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        fecha_convertida = pd.to_datetime(texto, format='%Y:%m:%d-%H:%M:%S', errors='coerce')
                    return fecha_convertida
                v['fecha'] = v['fecha'].apply(convertir_fecha)
                v['fecha'] = v['fecha'] 
                end_time = time.time()
                execution_time = end_time - start_time
                print("\n fecha inicial: ", v.fecha.min(),
                      "\n fecha final: ", v.fecha.max(),execution_time)
            #Variables
            m=list(v.columns)
            m.remove("fecha")
            for i in range(len(m)):
                v[m[i]] = pd.to_numeric(v[m[i]], errors='coerce')
                v[m[i]] = v[m[i]].astype(float)
            v["Rain_Rate"]=v['Rain_Rate']*0.2/60. #in units of cm/hr
            if Estacion in ["caucasia","santafe"]:
                v['Bar'] = (v['Bar']/1000.*(3386.389/100.0))
            else:
                v['Bar'] = (v['Bar']/1000.*(3386.389/100.0))
            v['Out_Temp'] = ( v['Out_Temp']/10. - 32.) * (5.0/9.0)
            v.loc[(v['Out_Temp'] > 50) | (v['Out_Temp'] < -15)] = np.nan
            #Procesamiento final
            v.sort_values("fecha")
            v = v.dropna(subset=['fecha'])
            diferencia = v['fecha'].diff()
            frecuencia_comun = diferencia.mode()[0]
            v = v.resample(frecuencia_comun, on='fecha').mean()
            v=v.reset_index()
            v.to_csv(path_out1+estacion+".csv")
            #Visualizacion
            graficoST(v,estacion)
            print("\n #######################################")
            #####################################################
############################## FIN CODIGO #####################################
